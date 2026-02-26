"""
Bangalore Vendor Data Collector — powered by SerpAPI (Free Tier)
=================================================================
Re-run safe: saves a NEW dated file each run + skips duplicate vendors.

Setup:
    1. Sign up FREE at https://serpapi.com  (no credit card needed)
    2. Copy your API key from https://serpapi.com/manage-api-key
    3. Paste it in SERPAPI_KEY below
    4. Install dependencies:
           pip install google-search-results pandas openpyxl phonenumbers

Run:
    python bangalore_vendor_scraper.py

Each run creates: Bangalore_Vendors_26-Feb-2026.xlsx  (date auto-updates)
Master file:      Bangalore_Vendors_Master_List.xlsx  (all runs combined, no duplicates)
"""

import time
import re
import os
import logging
from datetime import datetime
from typing import Optional

import pandas as pd
import phonenumbers
from serpapi import GoogleSearch
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# CONFIGURATION  <- Edit these values
# ─────────────────────────────────────────────
SERPAPI_KEY = "906424210fcf5e7285d2636fbe2fca84a3f901815e827390979042e9178c9849"   # https://serpapi.com/manage-api-key

BANGALORE_COORDS = "12.9716,77.5946"
SEARCH_RADIUS_METERS = 50000

VENDOR_CATEGORIES = [
    "Event Caterers Bangalore",
    "Tent House Bangalore",
    "Sound System Vendors Bangalore",
    "Wedding Decorators Bangalore",
    "Event Photographers Bangalore",
    "Florists Bangalore",
    "Wedding Venues Bangalore",
    "DJ Services Bangalore",
    "Event Planners Bangalore",
    "Lighting Equipment Rental Bangalore",
]

# File names — dated file changes every run, master file accumulates everything
TODAY = datetime.now().strftime("%d-%b-%Y")
DATED_FILE  = f"Bangalore_Vendors_{TODAY}.xlsx"
MASTER_FILE = "Bangalore_Vendors_Master_List.xlsx"

DELAY_BETWEEN_REQUESTS = 1.5

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)s  %(message)s")
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# PHONE NUMBER VALIDATION
# ─────────────────────────────────────────────
def validate_phone(raw: Optional[str]) -> Optional[str]:
    """Returns E.164 Indian phone number or None if invalid."""
    if not raw:
        return None
    cleaned = re.sub(r"[^\d+]", "", raw)
    for candidate in [cleaned, f"+91{cleaned.lstrip('0')}"]:
        try:
            parsed = phonenumbers.parse(candidate, "IN")
            if phonenumbers.is_valid_number(parsed):
                t = phonenumbers.number_type(parsed)
                valid_types = {
                    phonenumbers.PhoneNumberType.MOBILE,
                    phonenumbers.PhoneNumberType.FIXED_LINE,
                    phonenumbers.PhoneNumberType.FIXED_LINE_OR_MOBILE,
                }
                if t in valid_types:
                    return phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
        except phonenumbers.NumberParseException:
            continue
    return None


# ─────────────────────────────────────────────
# SERPAPI DATA FETCHING
# ─────────────────────────────────────────────
def search_category(category: str) -> list:
    """Fetches up to 60 results (3 pages x 20) for one category."""
    all_results = []

    for start in [0, 20, 40]:
        params = {
            "engine": "google_maps",
            "q": category,
            "ll": f"@{BANGALORE_COORDS},14z",
            "type": "search",
            "start": start,
            "api_key": SERPAPI_KEY,
        }
        try:
            search = GoogleSearch(params)
            results = search.get_dict()
            local_results = results.get("local_results", [])

            if not local_results:
                log.info(f"  No more results at page offset {start}")
                break

            all_results.extend(local_results)
            log.info(f"  Page offset {start}: {len(local_results)} results")
            time.sleep(DELAY_BETWEEN_REQUESTS)

        except Exception as e:
            log.warning(f"  SerpAPI error at offset {start}: {e}")
            break

    return all_results


# ─────────────────────────────────────────────
# RECORD EXTRACTION
# ─────────────────────────────────────────────
def extract_record(place: dict, category: str) -> Optional[dict]:
    """Parses one SerpAPI result. Returns None for permanently closed."""
    if place.get("permanently_closed") or "permanently closed" in str(place.get("status", "")).lower():
        return None

    raw_phone = place.get("phone")
    validated_phone = validate_phone(raw_phone)
    website = place.get("website") or place.get("links", {}).get("website", "N/A")

    return {
        "Category":             category.replace(" Bangalore", "").strip(),
        "Business Name":        place.get("title", "N/A"),
        "Phone Number (E.164)": validated_phone or "Not Available",
        "Phone Valid":          "Yes" if validated_phone else "No",
        "Address":              place.get("address", "N/A"),
        "Rating":               place.get("rating", "N/A"),
        "Total Reviews":        place.get("reviews", 0),
        "Website":              website if website else "N/A",
        "Google Maps Link":     place.get("link", "N/A"),
        "Date Collected":       TODAY,
    }


# ─────────────────────────────────────────────
# DEDUPLICATION
# ─────────────────────────────────────────────
def load_existing_master() -> pd.DataFrame:
    """Loads existing master file if it exists, else returns empty DataFrame."""
    if os.path.exists(MASTER_FILE):
        try:
            df = pd.read_excel(MASTER_FILE, sheet_name="All Vendors")
            log.info(f"Loaded existing master file: {len(df)} existing vendors")
            return df
        except Exception as e:
            log.warning(f"Could not read master file: {e}. Starting fresh.")
    return pd.DataFrame()


def deduplicate(new_df: pd.DataFrame, existing_df: pd.DataFrame) -> tuple[pd.DataFrame, int, int]:
    """
    Removes vendors from new_df that already exist in existing_df.
    Duplicate = same Business Name + same Address.
    Returns (new_only_df, new_count, duplicate_count).
    """
    if existing_df.empty:
        return new_df, len(new_df), 0

    # Create a unique key from name + address (lowercase, stripped)
    def make_key(df):
        return (
            df["Business Name"].str.lower().str.strip()
            + "||"
            + df["Address"].str.lower().str.strip()
        )

    existing_keys = set(make_key(existing_df))
    new_keys = make_key(new_df)
    is_new = ~new_keys.isin(existing_keys)

    new_only = new_df[is_new].reset_index(drop=True)
    duplicate_count = (~is_new).sum()

    return new_only, len(new_only), duplicate_count


# ─────────────────────────────────────────────
# MAIN COLLECTION PIPELINE
# ─────────────────────────────────────────────
def collect_all_vendors() -> pd.DataFrame:
    all_records = []

    for category in VENDOR_CATEGORIES:
        log.info(f"\nSearching: {category}")
        places = search_category(category)
        valid = 0

        for place in places:
            record = extract_record(place, category)
            if record:
                all_records.append(record)
                valid += 1

        log.info(f"  {valid} valid vendors found")

    log.info(f"\nTotal vendors this run: {len(all_records)}")
    return pd.DataFrame(all_records)


# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────
def export_to_excel(df: pd.DataFrame, filepath: str, label: str) -> None:
    if df.empty:
        log.warning(f"No data to export for {label}.")
        return

    summary = (
        df.groupby("Category")
        .agg(
            Total_Vendors=("Business Name", "count"),
            Valid_Phones=("Phone Valid", lambda x: (x == "Yes").sum()),
            Avg_Rating=("Rating", lambda x: pd.to_numeric(x, errors="coerce").mean()),
            Total_Reviews=("Total Reviews", "sum"),
        )
        .reset_index()
    )
    summary.columns = ["Category", "Total Vendors", "Valid Phones", "Avg Rating", "Total Reviews"]

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Vendors", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

    wb = load_workbook(filepath)
    _format_vendors_sheet(wb["All Vendors"])
    _format_summary_sheet(wb["Summary"], len(summary))
    wb.save(filepath)

    log.info(f"Saved {label}: {filepath}  ({len(df)} vendors)")


# ── Styling helpers ────────────────────────────────────────────────────────────
def _thin_border(color="D0D0D0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_header(cell, bg="1F3864"):
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border("FFFFFF")

def _apply_data(cell, row_idx):
    bg = "EBF0FA" if row_idx % 2 == 0 else "FFFFFF"
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _thin_border()

def _format_vendors_sheet(ws):
    widths = {"A": 22, "B": 35, "C": 20, "D": 12, "E": 45, "F": 9, "G": 14, "H": 32, "I": 35, "J": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        _apply_header(cell)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_idx].height = 20
        for cell in row:
            _apply_data(cell, row_idx)
            if cell.column == 4:
                if cell.value == "Yes":
                    cell.font = Font(name="Arial", size=10, color="006100", bold=True)
                    cell.fill = PatternFill("solid", start_color="C6EFCE")
                elif cell.value == "No":
                    cell.font = Font(name="Arial", size=10, color="9C0006", bold=True)
                    cell.fill = PatternFill("solid", start_color="FFC7CE")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    last = ws.max_row + 2
    for col, val in [(1, "TOTAL VENDORS"), (2, f"=COUNTA(B2:B{ws.max_row - 1})")]:
        c = ws.cell(row=last, column=col, value=val)
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F3864")
        c.alignment = Alignment(horizontal="center")

def _format_summary_sheet(ws, data_rows):
    for col, w in [("A", 28), ("B", 16), ("C", 14), ("D", 12), ("E", 16)]:
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        _apply_header(cell)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_idx].height = 20
        for cell in row:
            _apply_data(cell, row_idx)
            if cell.column == 4 and isinstance(cell.value, float):
                cell.number_format = "0.00"

    end = data_rows + 1
    last = end + 2
    for col, val in [
        (1, "GRAND TOTAL"),
        (2, f"=SUM(B2:B{end})"),
        (3, f"=SUM(C2:C{end})"),
        (4, f"=IFERROR(AVERAGE(D2:D{end}),0)"),
        (5, f"=SUM(E2:E{end})"),
    ]:
        c = ws.cell(row=last, column=col, value=val)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="2E75B6")
        c.alignment = Alignment(horizontal="center")
        if col == 4:
            c.number_format = "0.00"

    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────
def main():
    if SERPAPI_KEY == "YOUR_SERPAPI_KEY_HERE":
        raise ValueError(
            "\nPlease set your SerpAPI key.\n"
            "Sign up free at: https://serpapi.com\n"
            "Paste your key into SERPAPI_KEY in this file."
        )

    log.info("=" * 55)
    log.info("  Bangalore Vendor Collector — SerpAPI")
    log.info(f"  Run date : {TODAY}")
    log.info(f"  Categories: {len(VENDOR_CATEGORIES)}")
    log.info("=" * 55)

    # Step 1 — Collect fresh data from SerpAPI
    new_df = collect_all_vendors()

    if new_df.empty:
        log.warning("No vendors collected. Check your API key.")
        return

    # Step 2 — Load existing master to check for duplicates
    existing_df = load_existing_master()
    new_only_df, new_count, dup_count = deduplicate(new_df, existing_df)

    log.info(f"\nDeduplication Results:")
    log.info(f"  This run total   : {len(new_df)}")
    log.info(f"  Already in master: {dup_count}  (skipped)")
    log.info(f"  Genuinely new    : {new_count}  (added)")

    # Step 3 — Save DATED file (only this run's new vendors)
    if not new_only_df.empty:
        export_to_excel(new_only_df, DATED_FILE, "Dated file (new vendors only)")
    else:
        log.info(f"\nNo new vendors found — dated file not created (nothing new to save)")

    # Step 4 — Update MASTER file (all runs combined)
    if not existing_df.empty and not new_only_df.empty:
        combined_df = pd.concat([existing_df, new_only_df], ignore_index=True)
    elif not existing_df.empty:
        combined_df = existing_df
    else:
        combined_df = new_only_df

    export_to_excel(combined_df, MASTER_FILE, "Master file (all time)")

    log.info("\n" + "=" * 55)
    log.info(f"  DONE!")
    log.info(f"  Dated file : {DATED_FILE}  ({new_count} new vendors)")
    log.info(f"  Master file: {MASTER_FILE}  ({len(combined_df)} total vendors)")
    log.info("=" * 55)


if __name__ == "__main__":
    main()
